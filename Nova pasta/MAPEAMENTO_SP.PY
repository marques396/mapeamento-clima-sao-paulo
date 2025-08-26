from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import os.path
import tkinter as tk
from tkinter import ttk, messagebox
import subprocess # Módulo para abrir a pasta

class ClimaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Previsão do Tempo - São Paulo")
        self.root.geometry("400x350")
        
        self.style = ttk.Style()
        self.style.configure('TButton', font=('Arial', 12), padding=10)
        self.style.configure('TLabel', font=('Arial', 12), background='#f0f0f0')
        
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.pack(expand=True, fill=tk.BOTH)
        
        self.label_titulo = ttk.Label(self.main_frame, text="Previsão do Tempo", font=('Arial', 16, 'bold'))
        self.label_titulo.pack(pady=10)

        self.label_cidade = ttk.Label(self.main_frame, text="Digite a cidade:")
        self.label_cidade.pack(pady=(0, 5))
        self.entry_cidade = ttk.Entry(self.main_frame, width=30, font=('Arial', 12))
        self.entry_cidade.pack(pady=(0, 10))
        self.entry_cidade.insert(0, "São Paulo")

        self.btn_buscar = ttk.Button(self.main_frame, text="Buscar Previsão", command=self.executar_coleta)
        self.btn_buscar.pack(pady=(0, 20))
        
        self.frame_resultados = ttk.LabelFrame(self.main_frame, text="Resultados", padding=10)
        self.frame_resultados.pack(fill=tk.BOTH, expand=True)
        
        self.label_data = ttk.Label(self.frame_resultados, text="Data: -")
        self.label_data.pack(anchor='w')
        
        self.label_temp = ttk.Label(self.frame_resultados, text="Temperatura: -°C")
        self.label_temp.pack(anchor='w')
        
        self.label_umid = ttk.Label(self.frame_resultados, text="Umidade: -%")
        self.label_umid.pack(anchor='w')
        
        self.label_status = ttk.Label(self.frame_resultados, text="Condição: -")
        self.label_status.pack(anchor='w')
        
        self.label_status_oper = ttk.Label(self.main_frame, text="", foreground='blue')
        self.label_status_oper.pack(pady=5)
    
    def executar_coleta(self):
        try:
            self.btn_buscar['state'] = tk.DISABLED
            self.label_status_oper['text'] = "Conectando ao Google..."
            self.root.update()
            
            driver = self.iniciar_chrome()
            if driver:
                self.label_status_oper['text'] = "Coletando dados climáticos..."
                self.root.update()
                
                cidade = self.entry_cidade.get()
                dados = self.coletar_dados(driver, cidade)
                
                if dados:
                    self.atualizar_interface(dados)
                    self.salvar_dados_excel(dados, cidade)
                    self.label_status_oper['text'] = "Dados coletados e salvos com sucesso!"
                    messagebox.showinfo("Sucesso", "Dados coletados e salvos com sucesso!")
                    
                    # ADICIONADO: Abrir a pasta após salvar o arquivo
                    self.abrir_pasta_com_excel()

                else:
                    self.label_status_oper['text'] = "Falha ao coletar dados"
                    messagebox.showerror("Erro", "Não foi possível obter os dados climáticos")
            else:
                self.label_status_oper['text'] = "Falha ao iniciar navegador"
                messagebox.showerror("Erro", "Falha ao iniciar o navegador")
        
        except Exception as e:
            self.label_status_oper['text'] = f"Erro: {str(e)}"
            messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")
        finally:
            self.btn_buscar['state'] = tk.NORMAL
            self.root.update()
            if 'driver' in locals() and driver:
                driver.quit()
                os.system("taskkill /f /im chromedriver.exe /t >nul 2>&1")
                os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    
    def atualizar_interface(self, dados):
        self.label_data['text'] = f"Data: {dados['data']}"
        self.label_temp['text'] = f"Temperatura: {dados['temperatura']}°C"
        self.label_umid['text'] = f"Umidade: {dados['umidade']}%"
        self.label_status['text'] = f"Condição: {dados['status']}"
    
    def iniciar_chrome(self):
        try:
            chrome_options = Options()
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--log-level=3")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)

            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            return driver
            
        except Exception as e:
            print(f"Erro ao iniciar Chrome: {str(e)}")
            return None

    def coletar_dados(self, driver, cidade):
        try:
            query = f"clima {cidade}"
            driver.get(f'https://www.google.com.br/search?q={query}')
            
            wait = WebDriverWait(driver, 10)
            
            temp = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "span#wob_tm"))
            ).text
            
            umid = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "span#wob_hm"))
            ).text.replace('%', '')
            
            status = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "span#wob_dc"))
            ).text
            
            dados = {
                'data': time.strftime("%d/%m/%Y %H:%M"),
                'temperatura': temp,
                'umidade': umid,
                'status': status
            }
            
            return dados
            
        except Exception as e:
            print(f"Erro ao coletar dados: {str(e)}")
            return None

    def salvar_dados_excel(self, dados, cidade):
        try:
            pasta_destino = "DadosClima"
            nome_arquivo = "clima.xlsx"
            
            caminho_pasta = os.path.join(os.path.dirname(os.path.abspath(__file__)), pasta_destino)

            if not os.path.exists(caminho_pasta):
                os.makedirs(caminho_pasta)

            caminho_completo_arquivo = os.path.join(caminho_pasta, nome_arquivo)
            
            if os.path.exists(caminho_completo_arquivo):
                wb = load_workbook(caminho_completo_arquivo)
                sheet = wb.active
            else:
                wb = Workbook()
                sheet = wb.active
                sheet['A1'] = 'Cidade'
                sheet['B1'] = 'Data'
                sheet['C1'] = 'Temp (°C)'
                sheet['D1'] = 'Umidade (%)'
                sheet['E1'] = 'Status'
                
                for cell in sheet["1:1"]:
                    cell.font = Font(bold=True)
            
            nova_linha = [
                cidade, 
                dados['data'],
                dados['temperatura'],
                dados['umidade'],
                dados['status']
            ]
            sheet.append(nova_linha)
            
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(caminho_completo_arquivo)
            print(f"Dados salvos com sucesso em {caminho_completo_arquivo}")
            
        except Exception as e:
            print(f"Erro ao salvar dados no Excel: {str(e)}")
            raise
    
    def abrir_pasta_com_excel(self):
        try:
            pasta_destino = "DadosClima"
            caminho_pasta = os.path.join(os.path.dirname(os.path.abspath(__file__)), pasta_destino)
            
            if os.path.exists(caminho_pasta):
                if os.name == 'nt':  # Verifica se o sistema é Windows
                    os.startfile(caminho_pasta)
                elif os.name == 'posix': # Verifica se o sistema é Linux ou macOS
                    subprocess.Popen(['xdg-open', caminho_pasta]) # Linux
                    # Para macOS, você pode usar `subprocess.Popen(['open', caminho_pasta])`
            else:
                messagebox.showerror("Erro", f"A pasta '{pasta_destino}' não foi encontrada.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a pasta: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ClimaApp(root)
    root.mainloop()